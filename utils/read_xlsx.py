from django.db import connection
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from .make_defaults import save_model
from texts.models import Variable,Session,Response,Inputtype,Person,Text,Question
from texts.models import Transcriber
from . import extract_text

def open_question_file():
	t = open('../questions.txt').read().split('\n')
	t = [x for x in t if x]
	o = []
	for x in t:
		question = int(x.split('|')[0].split('Q1 ')[0].strip(' Q'))
		description = x.split('| ')[1].split(',')[0]
		condition = x.split(',')[-1].strip(' ')
		if condition == 'Beide condities':condition = 'Closed question'
		elif condition == 'Audioconditie':condition = 'Audio'
		elif condition == 'Tekstconditie':condition = 'text'	
		else: raise ValueError('unknown',condition)
		o.append([question,description,condition])
	return o

def read_questions_in_database():
	'''reads in questions into the database.'''
	questions_text = open_question_file()
	for qt in questions_text:
		number, description, condition = qt
		q = get_question(number)
		q.description = description
		q.title = description
		q.condition = condition
		q.save()

def link_response_to_session():
	'''links all responses in the database to a session if not linked yet.
	all sessions with an audio response will be linked to a session
	all other session are the sessions with a keyboard response
	'''
	for s in Session.objects.all():
		temp = Response.objects.filter(person__number= s.person)
		responses = temp.filter(response_date=s.session_date)
		for response in responses:
			if not response.session:
				t = response.text_set.all()[0].text
				i = response.question.column_index
				if eval(s.values)[i] == t:
					print('found match', 'linking session',s,'to:',response)
					response.session = s
					response.save()
			else:
				print('response',response,'already linked to session:')
				print(response.session)
				print('current session found:',s)

def _get_unlinked_sessions():
	'''get those sessions that are not linked to response.
	these are the sessions that have a text input.
	'''
	sessions = Session.objects.all()
	unlinked_sessions = []
	for session in sessions:
		if not session.response_set.all(): unlinked_sessions.append(session)
	return unlinked_sessions

def read_text_input_responses_in_database():
	''' read the response provide via keyboard into the database.'''
	unlinked_sessions = _get_unlinked_sessions()
	keyboard= Inputtype.objects.get(name = 'keyboard')
	text_questions = extract_text.get_questions(condition='text')
	for session in unlinked_sessions:
		date = session.session_date
		for i,question in enumerate(text_questions):
			index = question.column_index
			text = eval(session.values)[index]
			if text == -77:continue
			person = get_person(session.person)
			response = _make_response(question,person,keyboard,'','',
				date,session.row_index + 100000*i)
			response.session = session
			response.save()
			make_text(text, None, response, input_type= keyboard)


def open_nidi_xlsx():
	return load_workbook('../NIDI-voice-recorded-interviews-audio-transcripts.xlsx')

def open_text_xlsx():
	return load_workbook('../Text - Audio Matching.xlsx')

def handle_time(date_cell,time_cell):
	'''converts a date and time cell into a datetime object
	date_cell should have the following formate: 2021-04-17
	time_cell should have the following formate: 15:34:46
	'''
	s = date_cell + ' ' + time_cell
	return datetime.strftime(s,'%Y-%m-%d %H:%M:%S')

def read_in_variables():
	wb = open_nidi_xlsx()
	sheet = wb['Variables']
	for i,line in enumerate(list(sheet.values)[1:]):
		name, title, value = line[:3] 
		if not name: break
		if not title: title = ''
		if not value: value = ''
		d = {'name':name,'title':title,'value':value,'column_index':i}
		save_model(Variable,d,'name')

def session_header():
	wb = open_nidi_xlsx()
	sheet = wb['Response Data']
	return list(sheet.values)[0]
		
def read_in_session():
	wb = open_nidi_xlsx()
	sheet = wb['Response Data']
	for i,line in enumerate(list(sheet.values)[1:]):
		line = list(line)
		session_date = line[0]
		line[0] = str(line[0])
		duration = line[1]
		values = str(line)
		d = {'session_date':session_date,'values':values,'row_index':i}
		d.update({'duration':duration})
		save_model(Session,d,'row_index')

def make_audio_fn_dict(wb = None):
	if not wb: wb = open_text_xlsx()
	sheet = wb['audio_for_review']
	d = {}
	for i,line in enumerate(list(sheet.values)[1:]):
		d[line[0]] = line[:6]
	return d
	
def _add_date_and_time(date, time):
	d, t = date, time
	print('date',d,'time',t)
	return datetime(d.year,d.month,d.day,t.hour,t.minute,t.second)

def get_person(person_number):
	person_number = int(person_number)
	instance = save_model(Person,{'number':person_number},'number')
	return instance

def get_question(question_number):
	question_number = int(question_number)
	instance = save_model(Question,{'number':question_number},'number')
	return instance

def get_transcriber(name, human = False):
	instance = save_model(Transcriber,{'name':name,'human':human},'name')
	return instance

def make_text(text, transcriber, response, input_type= None):
	'''makes a new text instance based on a text and a transcriber.'''
	d ={'text':text,'transcriber':transcriber,'response':response}
	if input_type:d.update({'input_type':input_type})
	instance = save_model(Text,d)
	return instance
	
def _make_texts(tqf,tcd,toh,tps, response):
	'''special function to make texts instances form the 4 asr systems used
	to decode the audio.
	'''
	qf = Transcriber.objects.get(name ='questfox')
	cd = Transcriber.objects.get(name ='conversational_dialogues')
	oh = Transcriber.objects.get(name ='oral_history')
	ps = Transcriber.objects.get(name ='parliamentary_speeches')
	scribes = [qf,cd,oh,ps]
	texts = [tqf,tcd,toh,tps]
	return [make_text(t,s,response) for t,s in zip(texts,scribes)]
		
def _make_response(question,person,input_type,audio_filename,audio_quality,
	response_date,row_index):
	print('making response:',question,person)
	d = {'question':question,'person':person,'input_type':input_type}
	d.update({'audio_filename':audio_filename,'audio_quality':audio_quality})
	d.update({'response_date':response_date,'row_index':row_index})
	instance = save_model(Response,d,'row_index')
	return instance

def _line_ok(line):
	return line[0] != None


def read_in_text_audio_matching(clean_db = False):
	if clean_db: 
		Response.objects.all().delete()
		Text.objects.all().delete()
		connection.cursor().execute("VACUUM")
	qf = Transcriber.objects.get(name ='questfox')
	speech = Inputtype.objects.get(name = 'speech')
	wb = open_text_xlsx()
	audio_fn_dict = make_audio_fn_dict(wb = wb)
	sheet = wb['Matched Text Entries']
	for i,line in enumerate(list(sheet.values)[1:]):
		if not _line_ok(line): continue
		print('line',line)
		response_date = _add_date_and_time(line[0],line[1])
		person = get_person(line[4])
		question = get_question(line[7])
		audio_fn = line[8]
		audio_quality = ''
		response = _make_response(question,person,speech,audio_fn,audio_quality,
			response_date,i)
		if audio_fn in audio_fn_dict.keys():
			_, tqf, tcd, toh, tps, audio_quality = audio_fn_dict[audio_fn]
			texts = _make_texts(tqf,tcd,toh,tps,response)
		else:
			texts = [make_text(line[5],qf, response)]
			print('could not find:',audio_fn)

def add_manual_transcriptions():
	manual_transcriptions = open('../manual_text').read().split('\n')
	manual = Transcriber.objects.get(name='manual transcription')
	for line in manual_transcriptions:
		if not line or len(line.split('\t')) != 2:
			print(line)
			continue
		audio_filename, text = line.split('\t')	
		response = Response.objects.get(audio_filename=audio_filename)
		make_text(text,manual,response)


			
	
	
