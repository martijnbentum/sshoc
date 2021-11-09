from django.apps import apps
from transformers import pipeline
import pickle
from openpyxl import Workbook

def download_pipeline():
	'''download pipeline from huggingface.'''
	return pipeline('sentiment-analysis', 
		model='wietsedv/bert-base-dutch-cased-finetuned-sentiment')

def load_pipeline():
	'''opens a pickled pipeline locally.'''
	fin = open('sentiment_bertje','rb')
	p = pickle.load(fin)
	return p

def apply_pipeline(t, pipeline = None):
	if pipeline == None: pipeline = load_pipeline()
	return pipeline(t)

def text2label(text, pipeline = None):
	o = apply_pipeline(text.text, pipeline)
	return [text] + list(o[0].values())

def texts2label(texts, pipeline = None, label = ''):
	o = apply_pipeline([t.text for t in texts], pipeline)
	if label:
		return [[label, t.text] + list(v.values()) for t,v in zip(texts,o)]
	return [[t.text] + list(v.values()) for t,v in zip(texts,o)]
	

def _o2sheet(o,sheet,label = ''):
	for row,line in enumerate(o):
		for column,value in enumerate(line):
			sheet.cell(row=row+1,column = column+1,value = value)
		if label: sheet.cell(row=row+1, column = column +1, value = label)

	
def make_xlsx(questions = [15,20,29], filename = 'sentiment.xlsx'):
	p = load_pipeline()
	Question = apps.get_model('texts','question')
	wb = Workbook()
	for n in questions:
		q = Question.objects.get(number = n)
		o = texts2label(q.texts(), p)
		o = sorted(o, key=lambda x: x[1])
		sheet = wb.create_sheet(str(n))
		_o2sheet(o,sheet,q.description)
	del wb['Sheet']
	if filename:
		if not filename.endswith('.xlsx'): filename += '.xslx'
		wb.save(filename)
	return wb
